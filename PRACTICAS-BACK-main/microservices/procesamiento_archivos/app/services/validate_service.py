from fastapi import UploadFile, Form
import openpyxl
from io import BytesIO
from datetime import datetime
import httpx
import openpyxl.utils
from fastapi import Header
import unicodedata
import re
import holidays
from datetime import timedelta

CO_HOLIDAYS = holidays.CO()



TIPOS_PERMITIDOS = {
    "Auxilio de transporte": "SOLICITUDES.xlsx",
    "Descuento": "SOLICITUDES.xlsx",
    "Otros": "SOLICITUDES.xlsx",
    "Otro Si Definitivo": "SOLICITUDES.xlsx",
    "Horas Extra": "SOLICITUDES2.xlsx",
    "Otro Si Temporal": "SOLICITUDES3.xlsx",
    "Vacaciones": "SOLICITUDES4.xlsx",
}

MESES_ES = {
    "01": "enero", "02": "febrero", "03": "marzo", "04": "abril",
    "05": "mayo", "06": "junio", "07": "julio", "08": "agosto",
    "09": "septiembre", "10": "octubre", "11": "noviembre", "12": "diciembre"
}


def validar_horas_extra(fila, row_idx, campos_obligatorios, errores):
    fila_valida = True

    concepto_idx = campos_obligatorios.get("CONCEPTO")
    codigo_idx = campos_obligatorios.get("CON_CODIGO")
    unidad_idx = campos_obligatorios.get("UNIDADES")

    # ✅ Extraer valores
    concepto = str(fila[concepto_idx].value).strip() if concepto_idx is not None else ""
    codigo = str(fila[codigo_idx].value).strip() if codigo_idx is not None else ""
    unidad = str(fila[unidad_idx].value).strip() if unidad_idx is not None else ""

    concepto_valido_map = {
        "Domingo Sin Compensatorio Diurno": "75",
        "Domingo Sin Compensatorio Nocturno": "110",
        "Dominical Con Compensatorio Diurno": "66",
        "Dominical Con Compensatorio Nocturno": "78",
        "Festivo Sin Compensatorio Diurno": "75",
        "Festivo Sin Compensatorio Nocturno": "110",
        "Hora extra Diurna": "55",
        "Hora Extra Diurna": "55",  # Variación de mayúsculas
        "Hora extra Nocturna": "60",
        "Hora Extra Nocturna": "60",  # Variación de mayúsculas
        "Hora Extra festiva diurna": "65",  # Festivo diurno
        "Hora Extra festiva nocturna": "110",  # Festivo nocturno (por si se usa)
        "Recargo Nocturno 35%": "45",
    }

    # ✅ Validar CONCEPTO
    if concepto not in concepto_valido_map:
        errores.append(
            f"❌ Fila {row_idx}, Columna {concepto_idx + 1} (CONCEPTO): \"{concepto}\" no es válido. Verifica que esté escrito correctamente según las opciones disponibles."
        )
        fila_valida = False
    else:
        # ✅ Validar que el código corresponde al concepto
        codigo_esperado = concepto_valido_map[concepto]
        if codigo != codigo_esperado:
            errores.append(
                f"❌ Fila {row_idx}, Columna {codigo_idx + 1} (CON_CODIGO): se esperaba \"{codigo_esperado}\" para el concepto \"{concepto}\", pero llegó \"{codigo}\"."
            )
            fila_valida = False

    # ✅ Validar UNIDADES
    try:
        float(unidad)
    except Exception:
        errores.append(
            f"❌ Fila {row_idx}, Columna {unidad_idx + 1} (UNIDADES): debe ser un número válido (puede tener decimales). Valor ingresado: \"{unidad}\"."
        )
        fila_valida = False

    return fila_valida

def normalizar_texto(texto: str):
    texto = texto.lower()
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")  # quitar tildes
    return texto

def contiene_alguna_fecha_en_detalle(detalle: str, fecha_inicio: datetime, fecha_fin: datetime) -> bool:
    detalle_normalizado = normalizar_texto(detalle)

    def fecha_esta(fecha: datetime):
        dia = int(fecha.day)
        mes_num = fecha.strftime("%m")
        mes_nombre = MESES_ES[mes_num]
        anio = fecha.year

        patrones = [
            rf"\b{dia}\s*de\s*{mes_nombre}\s*de\s*{anio}\b",  # 7 de abril de 2025
            rf"\b{dia}\s*de\s*{mes_nombre}\b",
            rf"\b{dia}\s*{mes_nombre}\b",
            rf"\b{dia:02d}/{mes_num}/{anio}\b",
            rf"\b{dia}/{mes_num}/{anio}\b",
            rf"\b{dia:02d}-{mes_num}-{anio}\b",
            rf"\b{dia}-{mes_num}-{anio}\b",
            rf"\b{dia:02d}/{mes_num}\b",
            rf"\b{dia}/{mes_num}\b",
            rf"\b{dia:02d}-{mes_num}\b",
            rf"\b{dia}-{mes_num}\b",
        ]

        match_rango = re.search(r"\bdel\s+(\d{1,2})\s+al\s+(\d{1,2})\s+(?:de\s+)?([a-zA-ZñÑ]+)", detalle_normalizado)
        if match_rango:
            dia_inicio = int(match_rango.group(1))
            dia_fin = int(match_rango.group(2))
            mes_en_texto = match_rango.group(3).lower()
            if mes_en_texto == mes_nombre and dia_inicio <= dia <= dia_fin:
                return True

        return any(re.search(p, detalle_normalizado) for p in patrones)

    return fecha_esta(fecha_inicio) or fecha_esta(fecha_fin)


def validar_otro_si_temporal(fila, row_idx, campos_obligatorios, errores):
    fila_valida = True
    
    #VALIDACIÓN EN CAMPOS JORNADAS
    jornadas_validas = {"JORNADA 33%", "JORNADA 75%", "JORNADA 50%", "JORNADA 100%"}
    jornada_empleado_idx = campos_obligatorios.get("JORNADA EMPLEADO")
    jornada_otro_si_idx = campos_obligatorios.get("JORNADA OTRO SI TEMPORAL")
    
    # ✅ Extraer valores
    jornada_empleado = str(fila[jornada_empleado_idx].value).strip() if jornada_empleado_idx is not None else ""
    jornada_otro_si = str(fila[jornada_otro_si_idx].value).strip() if jornada_otro_si_idx is not None else ""
    
    # Relax validation: check if it STARTS with any valid prefix
    if not any(jornada_empleado.startswith(j) for j in jornadas_validas) and jornada_empleado: 
         # Si no está vacío y no coincide con niguna
         print(f"⚠️ Warning: Jornada empleado '{jornada_empleado}' no es estándar, pero se permite.")
        
    if not any(jornada_otro_si.startswith(j) for j in jornadas_validas) and jornada_otro_si:
         print(f"⚠️ Warning: Jornada otro si '{jornada_otro_si}' no es estándar, pero se permite.")
        
    #VALIDAR FECHAS
    fecha_inicio_idx = campos_obligatorios.get("FECHA INICIO")
    fecha_fin_idx = campos_obligatorios.get("FECHA FIN")
    
    # ✅ Extraer y normalizar fechas
    fecha_inicio_val = fila[fecha_inicio_idx].value if fecha_inicio_idx is not None else None
    fecha_fin_val = fila[fecha_fin_idx].value if fecha_fin_idx is not None else None

    fecha_inicio = normalizar_fecha(fecha_inicio_val)
    fecha_fin = normalizar_fecha(fecha_fin_val)
    
    if not isinstance(fecha_inicio, datetime):
        errores.append(
            f"❌ Fila {row_idx}, Columna {fecha_inicio_idx + 1} (FECHA INICIO): debe ser una fecha válida. Valor recibido: '{fecha_inicio_val}'"
        )
        fila_valida = False
    
    if not isinstance(fecha_fin, datetime):
        errores.append(
            f"❌ Fila {row_idx}, Columna {fecha_fin_idx + 1} (FECHA FIN): debe ser una fecha válida. Valor recibido: '{fecha_fin_val}'"
        )
        fila_valida = False
    
    #FECHAS IGUAL QUE EL DETALLE (SI LAS CONTIENE)
    detalle_idx = campos_obligatorios.get("DETALLE NOVEDAD")
    detalle = str(fila[detalle_idx].value).strip() if detalle_idx is not None and fila[detalle_idx].value else ""

    if detalle and isinstance(fecha_inicio, datetime) and isinstance(fecha_fin, datetime):
        detalle_normalizado = normalizar_texto(detalle)

        def fecha_en_detalle(fecha: datetime):
            dia = int(fecha.day)
            mes_num = fecha.strftime("%m")
            mes_nombre = MESES_ES[mes_num]
            anio = fecha.year

            patrones = [
                rf"\b{dia}\s*de\s*{mes_nombre}\s*de\s*{anio}\b",
                rf"\b{dia}\s*de\s*{mes_nombre}\b",
                rf"\b{dia}\s*{mes_nombre}\b",
                rf"\b{dia:02d}/{mes_num}/{anio}\b",
                rf"\b{dia}/{mes_num}/{anio}\b",
                rf"\b{dia:02d}-{mes_num}-{anio}\b",
                rf"\b{dia}-{mes_num}-{anio}\b",
                rf"\b{dia:02d}/{mes_num}\b",
                rf"\b{dia}/{mes_num}\b",
                rf"\b{dia:02d}-{mes_num}\b",
                rf"\b{dia}-{mes_num}\b",
            ]

            match_rango = re.search(r"\bdel\s+(\d{1,2})\s+al\s+(\d{1,2})\s+(?:de\s+)?([a-zA-ZñÑ]+)", detalle_normalizado)
            if match_rango:
                dia_inicio = int(match_rango.group(1))
                dia_fin = int(match_rango.group(2))
                mes_en_texto = match_rango.group(3).lower()
                if mes_en_texto == mes_nombre and dia_inicio <= dia <= dia_fin:
                    return True

            return any(re.search(p, detalle_normalizado) for p in patrones)

        if contiene_alguna_fecha_en_detalle(detalle, fecha_inicio, fecha_fin):
            faltan = []
            if not fecha_en_detalle(fecha_inicio):
                faltan.append(f"FECHA INICIO ({fecha_inicio.strftime('%d/%m/%Y')})")
            if not fecha_en_detalle(fecha_fin):
                faltan.append(f"FECHA FIN ({fecha_fin.strftime('%d/%m/%Y')})")

            if faltan:
                errores.append(
                    f"❌ Fila {row_idx}, Columna {detalle_idx + 1} (DETALLE NOVEDAD): Las fechas en las columnas 'FECHA INICIO' y 'FECHA FIN' no coinciden con lo que escribiste en el detalle. "
                    f"Verifica que las fechas en las columnas estén bien escritas. Fechas esperadas: {', '.join(faltan)}"
                )
                fila_valida = False


    #VALIDAR SALARIOS
    salario_actual_idx = campos_obligatorios.get("SALARIO ACTUAL")
    salario_otro_si_temp_idx = campos_obligatorios.get("SALARIO OTRO SI TEMPORAL")
    
    salario_actual = str(fila[salario_actual_idx].value).strip() if salario_actual_idx is not None else ""
    salario_otro_si_temp = str(fila[salario_otro_si_temp_idx].value).strip() if salario_otro_si_temp_idx is not None else ""
    
    try:
        float(salario_actual)
    except Exception:
        errores.append(
            f"❌ Fila {row_idx}, Columna {salario_actual_idx + 1} (SALARIO ACTUAL): debe ser numérico. Valor recibido: '{salario_actual}'"
        )
        fila_valida = False
        
    try:
        float(salario_otro_si_temp)
    except Exception:
        errores.append(
            f"❌ Fila {row_idx}, Columna {salario_otro_si_temp_idx + 1} (SALARIO OTRO SI TEMPORAL): debe ser numérico. Valor recibido: '{salario_otro_si_temp}'"
        )
        fila_valida = False
    
    #VALIDACION CONSECUTIVO FORMS
    consecutivo_idx = campos_obligatorios.get("CONSECUTIVO FORMS")
    consecutivo_val = str(fila[consecutivo_idx].value).strip() if consecutivo_idx is not None else ""
    
    if not consecutivo_val.upper().startswith("OT") and "caso atipico" not in consecutivo_val.lower():
        errores.append(
            f"❌ Fila {row_idx}, Columna {consecutivo_idx + 1} (CONSECUTIVO FORMS): debe comenzar con 'OT' o ser 'caso atipico'. Valor recibido: '{consecutivo_val}'"
        )
        fila_valida = False
    
    return fila_valida

#DIAS HABILES EN COLOMBIA
def es_dia_habil(date: datetime) -> bool:
    return date.weekday() < 5 and date.date() not in CO_HOLIDAYS

#CONTAR LOS DIAS HABILES EN COLOMBIA DE ACUERDO A LO INDICADO EN EL ARCHIVO
def contar_dias_habiles(inicio: datetime, fin: datetime) -> int:
    dias = 0
    actual = inicio
    while actual <= fin:
        if es_dia_habil(actual):
            dias += 1
        actual += timedelta(days=1)
    return dias

def validar_vacaciones(fila, row_idx, campos_obligatorios, errores):
    fila_valida = True

    dias_tomar_idx = campos_obligatorios.get("DIAS A TOMAR")
    fecha_inicio_disfrute_idx = campos_obligatorios.get("FECHA INICIO")
    fecha_fin_disfrute_idx = campos_obligatorios.get("FECHA FIN")

    if dias_tomar_idx is None or fecha_inicio_disfrute_idx is None or fecha_fin_disfrute_idx is None:
        errores.append(f"❌ Fila {row_idx}: Faltan columnas requeridas para vacaciones.")
        return False

    # Obtener valores
    try:
        dias_tomar = int(fila[dias_tomar_idx].value)
    except:
        errores.append(f"❌ Fila {row_idx}, Columna {dias_tomar_idx+1} (DIAS A TOMAR): debe ser un número entero.")
        fila_valida = False
        dias_tomar = None

    fecha_inicio = normalizar_fecha(fila[fecha_inicio_disfrute_idx].value)
    fecha_fin = normalizar_fecha(fila[fecha_fin_disfrute_idx].value)

    if not isinstance(fecha_inicio, datetime):
        errores.append(f"❌ Fila {row_idx}, Columna {fecha_inicio_disfrute_idx+1} (FECHA INICIO): debe ser una fecha válida.")
        fila_valida = False

    if not isinstance(fecha_fin, datetime):
        errores.append(f"❌ Fila {row_idx}, Columna {fecha_fin_disfrute_idx+1} (FECHA FIN): debe ser una fecha válida.")
        fila_valida = False

    if dias_tomar and fecha_inicio and fecha_fin:
        if fecha_inicio > fecha_fin:
            errores.append(f"❌ Fila {row_idx}: La FECHA INICIO no puede ser posterior a la FECHA FIN.")
            fila_valida = False
        else:
            dias_habiles = contar_dias_habiles(fecha_inicio, fecha_fin)
            if dias_habiles != dias_tomar:
                errores.append(
                    f"❌ Fila {row_idx}: Los días hábiles entre {fecha_inicio.strftime('%d/%m/%Y')} y {fecha_fin.strftime('%d/%m/%Y')} son {dias_habiles}, pero se indicó {dias_tomar}."
                )
                fila_valida = False

    return fila_valida
        

VALIDACIONES_ESPECIALES = {
    "Horas Extra": validar_horas_extra,
    "Otro Si Temporal": validar_otro_si_temporal,
    "Vacaciones": validar_vacaciones,
}

def normalizar_fecha(fecha_raw):
    if isinstance(fecha_raw, datetime):
        return fecha_raw.replace(hour=5, minute=0, second=0, microsecond=0)
    elif isinstance(fecha_raw, str):
        # Lista ampliada de formatos soportados
        formatos = [
            "%d/%m/%Y",       # 25/12/2025
            "%Y-%m-%d",       # 2025-12-25
            "%d-%m-%Y",       # 25-12-2025
            "%d/%m/%y",       # 25/12/25
            "%d-%m-%y",       # 25-12-25
            "%Y/%m/%d",       # 2025/12/25
            "%Y-%m-%d %H:%M:%S" # 2025-12-25 14:30:00
        ]
        
        fecha_limpia = fecha_raw.strip()
        # Reemplazar separadores comunes por guiones para unificar
        fecha_limpia = fecha_limpia.replace('/', '-').replace('.', '-').replace('\\', '-')
        
        for fmt in formatos:
            try:
                fecha = datetime.strptime(fecha_limpia, fmt)
                # Si el año es de 2 dígitos y se interpreta en el pasado lejano o futuro lejano, 
                # datetime lo maneja basado en pivot 2000 o similar, generalmente ok para este caso.
                return fecha.replace(hour=5, minute=0, second=0, microsecond=0)
            except ValueError:
                continue
    
    # Soporte para fechas numéricas de Excel (Serial Date)
    # Ej: 45915 -> 2025-09-16
    elif isinstance(fecha_raw, (int, float)):
        try:
            val = float(fecha_raw)
            # Solo procesar números razonables para fechas (ej: > 1000) para evitar falsos positivos con otros números
            if val > 1000:
                return openpyxl.utils.datetime.from_excel(val).replace(hour=5, minute=0, second=0, microsecond=0)
        except Exception:
            pass
    
    # Soporte para cadenas que son números (Excel date como string)
    elif isinstance(fecha_raw, str) and fecha_raw.replace('.', '', 1).isdigit():
        try:
             val = float(fecha_raw)
             if val > 1000:
                 return openpyxl.utils.datetime.from_excel(val).replace(hour=5, minute=0, second=0, microsecond=0)
        except Exception:
            pass

    return None


async def validar_duplicado_en_backend(cedula: str, fecha: datetime, tipo: str, nombre: str, jwt_token: str):
    url = "http://localhost:3000/novedad/validar-duplicado"

    fecha_str = fecha.strftime("%Y-%m-%d %H:%M:%S")
    params = {
        "cedula": cedula,
        "fecha": fecha_str,
        "tipo": tipo,
        "nombre": nombre
    }

    # 👇 Asegura que el token tenga el prefijo "Bearer "
    if jwt_token and not jwt_token.lower().startswith("bearer "):
        jwt_token = f"Bearer {jwt_token}"

    headers = {
        "Authorization": jwt_token
    }

    print(f"🔍 [MICROSERVICIO] Enviando request a BD: {params}")

    try:
        async with httpx.AsyncClient() as client:
            response = await client.get(url, params=params, headers=headers, timeout=10)
            print(f"🔍 [MICROSERVICIO] Status code: {response.status_code}")
            print(f"🔍 [MICROSERVICIO] Response text: {response.text}")
            
            if response.status_code == 200:
                data = response.json()
                print(f"🔍 [MICROSERVICIO] Response data: {data}")
                existe = data.get("existe", False)
                mensaje = data.get("mensaje", "")
                print(f"🔍 [MICROSERVICIO] ¿Existe duplicado? {existe}")
                return existe, mensaje
            else:
                print(f"❌ [MICROSERVICIO] Error HTTP: {response.status_code}")
                return False, f"⚠️ Error al consultar duplicado: código {response.status_code}"
    except Exception as e:
        print(f"❌ [MICROSERVICIO] Excepción: {e}")
        return False, f"⚠️ Excepción al validar duplicado: {e}"


async def validar_excel(
    file: UploadFile,
    tipo: str,
    titulo: str = Form(...),
    nombreUsuario: str = Form(...),
    nombreTienda: str = Form(...),
    authorization: str = Header(None), 
):
    
    print(f"📥 Archivo recibido: {file.filename}")
    print(f"📋 Tipo: {tipo}, Título: {titulo}, Usuario: {nombreUsuario}, Tienda: {nombreTienda}")
    
    errores = []
    cantidad_solicitudes = 0  # Contador de filas válidas

    if tipo not in TIPOS_PERMITIDOS:
        return {
            "valido": False,
            "errores": [f"❌ Tipo de solicitud '{tipo}' no es válido."]
        }

    content = await file.read()
    wb = openpyxl.load_workbook(filename=BytesIO(content), data_only=True)
    sheet = wb.active

    encabezados = [cell.value for cell in sheet[5]]
    encabezados_normalizados = [str(h).strip().upper() if h else "" for h in encabezados]
    print("📌 Encabezados detectados:", encabezados_normalizados)

    # Diccionario de cabeceras esperadas por tipo
    cabeceras_por_tipo = {
        "SOLICITUDES.xlsx": [
            "N", "FECHA DE REPORTE", "CEDULA", "NOMBRE (APELLIDOS-NOMBRES)",
            "CATEGORIA", "TIENDA", "QUIEN REPORTA LA NOVEDAD\n(Nombre Jefe GH)", "DETALLE NOVEDAD"
        ],
        "SOLICITUDES2.xlsx": [
            "N", "FECHA DE REPORTE", "CEDULA", "NOMBRE (APELLIDOS-NOMBRES)",
            "CATEGORIA", "TIENDA", "QUIEN REPORTA LA NOVEDAD\n(Nombre Jefe GH)", "DETALLE NOVEDAD",
            "CONCEPTO", "CON_CODIGO", "UNIDADES", "FECHA NOVEDAD"
        ],
        "SOLICITUDES3.xlsx": [
            "N", "FECHA DE REPORTE", "CEDULA", "NOMBRE (APELLIDOS-NOMBRES)",
            "CATEGORIA", "TIENDA", "QUIEN REPORTA LA NOVEDAD\n(Nombre Jefe GH)", "DETALLE NOVEDAD",
            "JORNADA EMPLEADO", "JORNADA OTRO SI TEMPORAL", "FECHA INICIO", "FECHA FIN",
            "SALARIO ACTUAL", "SALARIO OTRO SI TEMPORAL", "CONSECUTIVO FORMS"
        ],
        "SOLICITUDES4.xlsx": [
            "N", "FECHA DE REPORTE", "CEDULA", "NOMBRE (APELLIDOS-NOMBRES)",
            "CATEGORIA", "TIENDA", "QUIEN REPORTA LA NOVEDAD\n(Nombre Jefe GH)", "DETALLE NOVEDAD",
            "DIAS A TOMAR", "FECHA INICIO", "FECHA FIN"
        ]
    }
    
    CAMPOS_OBLIGATORIOS_POR_PLANTILLA = {
        "SOLICITUDES.xlsx": [
            "CEDULA", "NOMBRE (APELLIDOS-NOMBRES)", "DETALLE NOVEDAD"
        ],
        "SOLICITUDES2.xlsx": [
            "CEDULA", "NOMBRE (APELLIDOS-NOMBRES)", "DETALLE NOVEDAD",
            "CONCEPTO", "CON_CODIGO", "UNIDADES", "FECHA NOVEDAD"
        ],
        "SOLICITUDES3.xlsx": [
            "CEDULA", "NOMBRE (APELLIDOS-NOMBRES)", "DETALLE NOVEDAD",
            "JORNADA EMPLEADO", "JORNADA OTRO SI TEMPORAL", "FECHA INICIO", "FECHA FIN",
            "SALARIO ACTUAL", "SALARIO OTRO SI TEMPORAL", "CONSECUTIVO FORMS"
        ],
        "SOLICITUDES4.xlsx": [
            "CEDULA", "NOMBRE (APELLIDOS-NOMBRES)", "DETALLE NOVEDAD",
            "DIAS A TOMAR", "FECHA INICIO", "FECHA FIN"
        ]
    }

    plantilla_esperada = TIPOS_PERMITIDOS[tipo]
    
    # Construir los campos obligatorios directamente desde los headers detectados
    # Usar matching flexible por palabras clave
    campos_obligatorios = {}
    
    campos_esperados_map = CAMPOS_OBLIGATORIOS_POR_PLANTILLA[plantilla_esperada]
    
    # Para cada campo esperado, buscar en los headers recibidos
    for campo_esperado in campos_esperados_map:
        campo_upper = campo_esperado.upper()
        
        # Buscar coincidencia exacta primero
        if campo_upper in encabezados_normalizados:
            campos_obligatorios[campo_esperado] = encabezados_normalizados.index(campo_upper)
            continue
        
        # Si no hay coincidencia exacta, hacer matching fuzzy por palabra clave
        palabras_clave = {
            "CEDULA": ["CEDULA", "CC", "IDENTIFICACION", "DOCUMENTO"],
            "NOMBRE (APELLIDOS-NOMBRES)": ["NOMBRE", "APELLIDOS", "EMPLEADO", "COLABORADOR"],
            "DETALLE NOVEDAD": ["DETALLE", "NOVEDAD", "OBSERVACION", "DESCRIPCION"],
            "CONCEPTO": ["CONCEPTO"],
            "CON_CODIGO": ["CODIGO", "CON_CODIGO", "ID_CONCEPTO"],
            "UNIDADES": ["UNIDADES", "CANTIDAD", "HORAS"],
            "FECHA NOVEDAD": ["FECHA NOVEDAD", "FECHA_NOVEDAD", "DIA NOVEDAD"],
            "JORNADA EMPLEADO": ["JORNADA EMPLEADO", "JORNADA", "TURNO"],
            "JORNADA OTRO SI TEMPORAL": ["JORNADA OTRO SI", "NUEVA JORNADA", "CAMBIO TURNO"],
            "FECHA INICIO": ["FECHA INICIO", "INICIO", "DESDE", "F. INICIO"],
            "FECHA FIN": ["FECHA FIN", "FIN", "HASTA", "F. FIN", "TERMINACION"],
            "SALARIO ACTUAL": ["SALARIO ACTUAL", "SUELDO", "BASICO"],
            "SALARIO OTRO SI TEMPORAL": ["SALARIO OTRO SI", "NUEVO SALARIO", "NUEVO SUELDO"],
            "CONSECUTIVO FORMS": ["CONSECUTIVO", "FORMS", "OT"],
            "DIAS A TOMAR": ["DIAS", "TOMAR", "CANTIDAD DIAS", "NUMERO DIAS"],
        }
        
        print("🔍 [DEBUG] Buscando columnas obligatorias...")
        print(f"🔍 [DEBUG] Headers normalizados: {encabezados_normalizados}")
        
        # Buscar por palabra clave
        encontrado = False
        for campo_esperado in campos_esperados_map:
             print(f"🔍 [DEBUG] Buscando campo: {campo_esperado}")
             encontrado_campo = False
             
             # 1. Búsqueda exacta
             campo_upper = campo_esperado.upper()
             if campo_upper in encabezados_normalizados:
                 campos_obligatorios[campo_esperado] = encabezados_normalizados.index(campo_upper)
                 print(f"✅ [DEBUG] Encontrado EXACTO: {campo_esperado} en índice {campos_obligatorios[campo_esperado]}")
                 continue

             # 2. Búsqueda por palabras clave
             if campo_esperado in palabras_clave:
                for keyword in palabras_clave[campo_esperado]:
                    keyword_normalized = keyword.upper()
                    for idx, header in enumerate(encabezados_normalizados):
                        if keyword_normalized in header:
                            campos_obligatorios[campo_esperado] = idx
                            print(f"✅ [DEBUG] Encontrado por KEYWORD '{keyword}': {campo_esperado} en índice {idx} (Header: '{header}')")
                            encontrado_campo = True
                            break
                    if encontrado_campo:
                        break
             
             if not encontrado_campo and campo_esperado not in campos_obligatorios:
                 print(f"❌ [DEBUG] NO ENCONTRADO: {campo_esperado}")
    
    # Validar que encontramos todas las columnas necesarias
    faltantes = [f for f in campos_esperados_map if f not in campos_obligatorios]
    if faltantes:
        errores.append(f"❌ Faltan las siguientes columnas obligatorias: {faltantes}")
        errores.append(f"📄 Headers detectados: {encabezados_normalizados}")

    if errores:
        print("🛑 Errores por campos obligatorios:", errores)
        return {
            "valido": False,
            "errores": errores
        }

    # SET CONTROL DE DUPLICADOS EN LA MISMA PLANTILLA
    duplicados_cedula_fecha = set()
    
    # MEJORADO: Validar todos los registros contra la BD al inicio
    registros_para_validar = []
    
    # Primero recopilar todos los registros válidos
    for row_idx in range(6, sheet.max_row + 1):
        fila = sheet[row_idx]
        fila_visible = any(cell.value is not None for cell in fila)

        if not fila_visible:
            continue

        # Extraer datos principales
        cedula_idx = campos_obligatorios.get("CEDULA")
        fecha_idx = 1  # COLUMNA B --> FECHA
        
        cedula_val = str(fila[cedula_idx].value).strip() if cedula_idx is not None else ""
        fecha_val = fila[fecha_idx].value
        
        if isinstance(fecha_val, datetime):
            fecha_val = fecha_val.replace(hour=5, minute=0, second=0, microsecond=0)
        elif isinstance(fecha_val, str):
            try:
                fecha_val = datetime.strptime(fecha_val.strip(), "%d/%m/%Y").replace(hour=5, minute=0, second=0, microsecond=0)
            except Exception:
                fecha_val = None
        
        if cedula_val and fecha_val:
            registros_para_validar.append({
                'row_idx': row_idx,
                'cedula': cedula_val,
                'fecha': fecha_val,
                'tipo': tipo
            })

    # Validar duplicados en BD de forma masiva
    print(f"🔍 [MICROSERVICIO] Validando {len(registros_para_validar)} registros contra la BD...")
    duplicados_bd = set()
    
    for registro in registros_para_validar:
        print(f"🔍 [MICROSERVICIO] Validando registro: {registro}")
        
        jwt_token = str(authorization)
        
        existe_en_bd, mensaje_duplicado = await validar_duplicado_en_backend(
            registro['cedula'],
            registro['fecha'],
            registro['tipo'],
            "",  # nombre no es necesario para la validación de duplicados
            jwt_token 
        )
        
        print(f"🔍 [MICROSERVICIO] Resultado para {registro['cedula']}: existe={existe_en_bd}, mensaje={mensaje_duplicado}")
        
        if existe_en_bd:
            # Crear clave única para identificar el duplicado
            clave_duplicado = f"{registro['cedula']}-{registro['fecha'].strftime('%Y-%m-%d')}-{registro['tipo']}"
            duplicados_bd.add(clave_duplicado)
            print(f"⚠️ [MICROSERVICIO] Duplicado encontrado en BD: {clave_duplicado}")

    print(f"🔍 [MICROSERVICIO] Total duplicados encontrados en BD: {len(duplicados_bd)}")
    
    # Ahora procesar todas las filas
    for row_idx in range(6, sheet.max_row + 1):
        fila = sheet[row_idx]
        fila_visible = any(cell.value is not None for cell in fila)

        if not fila_visible:
            print(f"⚪ Fila {row_idx} vacía. Saltando...")
            continue

        print(f"\n🔍 Validando fila {row_idx}...")
        fila_valida = True

        # Validar campos obligatorios
        for campo, col_idx in campos_obligatorios.items():
            cell = fila[col_idx]
            valor = cell.value

            if valor is None or str(valor).strip() == "":
                mensaje_error = f"❌ Fila {row_idx}, Columna {col_idx + 1} ({campo}): VACÍA"
                print(mensaje_error)
                errores.append(mensaje_error)
                fila_valida = False
            else:
                print(f"✅ Fila {row_idx}, Columna {col_idx + 1} ({campo}): OK → '{valor}'")
            
                # VALIDACIONES EXTRA
                
                # 3. LONGITUD DE MENSAJE EN DETALLE NOVEDAD
                if campo == "DETALLE NOVEDAD":
                    texto = str(valor).strip()
                    longitud = len(texto)
                    if longitud < 15 or longitud > 500:
                        errores.append(
                            f"❌ Fila {row_idx}, Columna {col_idx + 1} (DETALLE NOVEDAD): debe tener entre 15 y 500 caracteres. Tiene {longitud}."
                        )
                        fila_valida = False
                
                # 4. SIN CARACTERES RAROS EN NOMBRE
                if campo == "NOMBRE (APELLIDOS-NOMBRES)":
                    texto = str(valor).strip()
                    if not all(char.isalpha() or char.isspace() or char in "áéíóúÁÉÍÓÚñÑ" for char in texto):
                        errores.append(
                            f"❌ Fila {row_idx}, Columna {col_idx + 1} (NOMBRE): contiene caracteres inválidos. Solo se permiten letras, tildes y espacios."
                        )
                        fila_valida = False
                
                # 5. CAMPO CEDULA SOLO NUMEROS
                if campo == "CEDULA":
                    texto = str(valor).strip()

                    if texto.lower() == "none" or texto == "":
                        errores.append(
                            f"❌ Fila {row_idx}, Columna {col_idx + 1} (CÉDULA): no puede estar vacía ni ser 'None'."
                        )
                        fila_valida = False
                    elif not texto.isdigit():
                        errores.append(
                            f"❌ Fila {row_idx}, Columna {col_idx + 1} (CÉDULA): debe contener solo números. Valor recibido: '{texto}'."
                        )
                        fila_valida = False


        # Validaciones automáticas por columnas generadas
        cell_A = fila[0]  # Columna A
        cell_B = fila[1]  # Columna B
        cell_E = fila[4]  # Columna E
        cell_F = fila[5]  # Columna F
        cell_G = fila[6]  # Columna G

        # A: Número secuencial
        numero_esperado = row_idx - 5
        # A: Número secuencial - RELAJADO A WARNING
        numero_esperado = row_idx - 5
        if cell_A.value != numero_esperado:
            print(f"⚠️ Advertencia Fila {row_idx}, columna A: Se esperaba el número '{numero_esperado}' y llegó '{cell_A.value}'. Se permite continuar.")
            # ERROR ELIMINADO PARA PERMITIR SALTOS
            # errores.append(f"❌ Fila {row_idx}, columna A: Se esperaba el número '{numero_esperado}' y llegó '{cell_A.value}'")
            # fila_valida = False

        # B: Fecha del día
        fecha_b = normalizar_fecha(cell_B.value)
        hoy = normalizar_fecha(datetime.now())
        if fecha_b != hoy:
             print(f"⚠️ Advertencia Fila {row_idx}, columna B: Se esperaba la fecha '{hoy.strftime('%d/%m/%Y')}' y llegó '{cell_B.value}'. Se permite continuar.")
             # ERROR ELIMINADO PARA PERMITIR FECHAS DIFERENTES
             # errores.append(f"❌ Fila {row_idx}, columna B: Se esperaba la fecha '{hoy.strftime('%d/%m/%Y')}' y llegó '{cell_B.value}'")
             # fila_valida = False

        # E: Título / CATEGORIA - RELAJADO A WARNING
        # Esta columna es CATEGORIA (dato de fila), no el tipo de novedad
        if str(cell_E.value).strip() != titulo.strip():
            print(f"⚠️ Advertencia Fila {row_idx}, columna E (CATEGORIA): Se esperaba '{titulo}' y llegó '{cell_E.value}'. Se permite continuar.")
            # ERROR ELIMINADO - No bloquear por CATEGORIA diferente
            # fila_valida = False

        # F: Tienda
        # F: Tienda
        if str(cell_F.value).strip() != nombreTienda.strip():
             print(f"⚠️ Advertencia Fila {row_idx}, columna F: Se esperaba la tienda '{nombreTienda}' y llegó '{cell_F.value}'. Se permite continuar.")
             # ERROR ELIMINADO PARA PERMITIR OTRAS TIENDAS
             # errores.append(f"❌ Fila {row_idx}, columna F: Se esperaba la tienda '{nombreTienda}' y llegó '{cell_F.value}'")
             # fila_valida = False

        # G: Jefe
        # G: Jefe
        if str(cell_G.value).strip() != nombreUsuario.strip():
             print(f"⚠️ Advertencia Fila {row_idx}, columna G: Se esperaba el nombre del jefe '{nombreUsuario}' y llegó '{cell_G.value}'. Se permite continuar.")
             # ERROR ELIMINADO PARA PERMITIR OTROS JEFES
             # errores.append(f"❌ Fila {row_idx}, columna G: Se esperaba el nombre del jefe '{nombreUsuario}' y llegó '{cell_G.value}'")
             # fila_valida = False
        
        # VALIDACIONES EXTRA
        
        # 2. No puede estar la misma persona el mismo día en la misma solicitud (dentro del archivo)
        cedula_idx = campos_obligatorios.get("CEDULA")
        fecha_idx = 1  # COLUMNA B --> FECHA
        
        cedula_val = str(fila[cedula_idx].value).strip() if cedula_idx is not None else ""
        fecha_val = fila[fecha_idx].value
        
        if isinstance(fecha_val, datetime):
            fecha_val = fecha_val.replace(hour=5, minute=0, second=0, microsecond=0)
        elif isinstance(fecha_val, str):
            try:
                fecha_val = datetime.strptime(fecha_val.strip(), "%d/%m/%Y").replace(hour=5, minute=0, second=0, microsecond=0)
            except Exception:
                fecha_val = None
                
        # Evitar validación de duplicados si la cédula es inválida
        if not cedula_val or cedula_val.lower() == "none":
            print(f"⚠️ Fila {row_idx}: Cédula vacía o inválida. Se omite validación de duplicados.")
        else:
            clave_archivo = f"{cedula_val}-{fecha_val}"
            if clave_archivo in duplicados_cedula_fecha:
                # RELAJADO A WARNING
                print(f"⚠️ Advertencia Fila {row_idx}: La persona con cédula {cedula_val} ya tiene una solicitud registrada el día {fecha_val.strftime('%d/%m/%Y')} en este archivo. Se permite.")
            else:
                duplicados_cedula_fecha.add(clave_archivo)
        
        # VALIDAR DUPLICADOS EN BD (usando los datos pre-validados)
        if fecha_val:  # Solo validar si la fecha es válida
            clave_bd = f"{cedula_val}-{fecha_val.strftime('%Y-%m-%d')}-{tipo}"
            print(f"🔍 [MICROSERVICIO] Verificando clave BD: {clave_bd}")
            print(f"🔍 [MICROSERVICIO] Duplicados BD encontrados: {duplicados_bd}")
            
            if clave_bd in duplicados_bd:
                mensaje_error = f"⚠️ Advertencia Fila {row_idx}: Ya existe una novedad en la base de datos con cédula {cedula_val}, fecha {fecha_val.strftime('%d/%m/%Y')} y tipo '{tipo}'. Se permite duplicado."
                print(f"⚠️ [MICROSERVICIO] {mensaje_error}")
                # ERROR ELIMINADO PARA PERMITIR DUPLICADOS
                # errores.append(mensaje_error)
                # fila_valida = False
            else:
                print(f"✅ [MICROSERVICIO] No hay duplicado en BD para: {clave_bd}")
        
        if tipo in VALIDACIONES_ESPECIALES and fila_valida:
            funcion_validadora = VALIDACIONES_ESPECIALES[tipo]
            resultado_especial = funcion_validadora(fila, row_idx, campos_obligatorios, errores)
            if not resultado_especial:
                print(f"⚠️ Fila {row_idx} falló la validación especial para tipo '{tipo}'")
                fila_valida = False
        
        if fila_valida:
            print(f"✅ Fila {row_idx} completa y válida.")
            cantidad_solicitudes += 1
        else:
            print(f"⚠️ Fila {row_idx} tiene errores.")

    if errores:
        print("\n🛑 Validación terminada con errores.")
        print(f"🛑 Total errores: {len(errores)}")
        return {
            "valido": False,
            "errores": errores,
            "tipoValidado": tipo,
            "cantidadSolicitudes": cantidad_solicitudes
        }

    print(f"\n✅ Validación exitosa. Total solicitudes válidas: {cantidad_solicitudes}")
    return {
        "valido": True,
        "esMasiva": True,
        "cantidadSolicitudes": cantidad_solicitudes,
        "tipoValidado": tipo
    }